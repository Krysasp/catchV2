#!/usr/bin/env python3
"""
Output writer module for OligoPlex.

Handles multi-sheet Excel output generation.
"""

import os
from typing import Dict, List, Any, Optional


def generate_renamed_consensus_fasta(cluster_taxonomy: Dict[str, Dict[str, str]],
                                     degapped_consensus: Dict[str, str],
                                     analysis_stats: Dict[str, Dict[str, float]],
                                     cluster_sequences: Dict[str, List[Dict]],
                                     output_path: str) -> int:
    """
    Generate FASTA file with renamed consensus sequences.
    
    Header format: >Ccluster_num|family|genus|num_sequences|avg_coverage_over_unambig
    
    Example: >C0011|Adenoviridae|Mastadenovirus|18|3.639
    
    Args:
        cluster_taxonomy: dict mapping cluster_num to taxonomy info
        degapped_consensus: dict mapping cluster_num to degapped consensus sequence
        analysis_stats: dict mapping cluster_num to analysis statistics
        cluster_sequences: dict mapping cluster_num to list of sequence info
        output_path: output FASTA path
    
    Returns:
        number of consensus sequences written
    """
    count = 0
    
    with open(output_path, 'w') as f:
        for cluster_num in sorted(degapped_consensus.keys(), key=int):
            tax_info = cluster_taxonomy.get(cluster_num, {})
            stats_info = analysis_stats.get(cluster_num, {})
            seq_info_list = cluster_sequences.get(cluster_num, [])
            
            # Get taxonomy info
            family = tax_info.get('family', 'Unknown').replace(' ', '_')
            genus = tax_info.get('genus', 'Unknown').replace(' ', '_')
            species = tax_info.get('species', 'Unknown').replace(' ', '_')
            
            # Get stats
            num_seqs = len(seq_info_list)
            avg_coverage = stats_info.get('avg_coverage_over_unambig', 0)
            
            # Format cluster number with leading zeros
            cluster_display = f"C{int(cluster_num):04d}"
            
            # Format header
            header = f">{cluster_display}|{family}|{genus}|{species}|{num_seqs}|{avg_coverage:.3f}"
            
            # Write sequence
            consensus_dict = degapped_consensus[cluster_num]
            consensus_seq = consensus_dict['sequence']
            f.write(f"{header}\n")
            f.write(f"{consensus_seq}\n")
            
            count += 1
    
    return count


def write_probe_mapping_excel(output_path: str,
                              analysis: Dict[str, Dict],
                              probe_details: Dict[str, Dict],
                              cluster_accessions: Dict[str, List[str]],
                              metadata_map: Dict[str, Dict],
                              adapter_seq: str = 'GTGGAGGTCGCTAGATGGTC') -> None:
    """
    Write comprehensive Excel output with multiple sheets.
    
    Sheets:
    1. ClusterSummary: Extended analysis TSV with taxonomy details (omitting 'Genome' column, adding 'Molecule_type')
    2. ProbeMapping: All probe mapping results with coordinates (excluding adapter portion)
    
    Args:
        output_path: output Excel file path
        analysis: dict mapping genome_id to analysis stats
        probe_details: dict mapping probe_id to probe details
        cluster_accessions: dict mapping cluster_num to list of accession IDs
        metadata_map: dict mapping accession_id to taxonomy info
        adapter_seq: adapter sequence to exclude from coordinates
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("[WARN] openpyxl not available, using csv fallback")
        _write_probe_mapping_csv(output_path, probe_details, analysis, cluster_accessions, metadata_map)
        return
    
    workbook = openpyxl.Workbook()
    
    # Sheet 1: Cluster Summary (extended analysis with taxonomy, omitting 'Genome' column, adding 'Molecule_type')
    cluster_sheet = workbook.active
    cluster_sheet.title = 'ClusterSummary'
    
    # Header row - based on analysis TSV but omit 'Genome', add 'Molecule_type'
    cluster_headers = [
        'Cluster_ID', 'Num_bases_covered', 'Frac_bases_covered', 'Frac_bases_covered_over_unambig',
        'Average_coverage_depth', 'Average_coverage_depth_over_unambig',
        'Family', 'Genus', 'Species', 'Molecule_type', 'Num_accessions'
    ]
    
    for col, header in enumerate(cluster_headers, 1):
        cell = cluster_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Get unique cluster numbers from analysis
    cluster_nums = set()
    for genome_id in analysis.keys():
        # Extract cluster number from genome_id (format: "final_consensus_sequences.fasta, genome N")
        if 'genome' in genome_id:
            try:
                cluster_num = int(genome_id.split(', genome ')[1])
                cluster_nums.add(cluster_num)
            except (ValueError, IndexError):
                pass
    
    # If no clusters found in analysis, use cluster_accessions
    if not cluster_nums:
        cluster_nums = set(cluster_accessions.keys())
    
    # Data rows
    for row_idx, cluster_num in enumerate(sorted(cluster_nums, key=int), 2):
        # Get analysis stats
        # Find matching genome_id
        genome_id = None
        for gid in analysis.keys():
            if f'genome {cluster_num}' in gid:
                genome_id = gid
                break
        
        stats = analysis.get(genome_id, {}) if genome_id else {}
        
        # Get taxonomy from first accession in cluster
        taxonomy = {}
        for acc in cluster_accessions.get(cluster_num, []):
            if acc in metadata_map:
                taxonomy = metadata_map[acc]
                break
        
        if not taxonomy:
            taxonomy = {'family': 'Unknown', 'genus': 'Unknown', 'species': 'Unknown', 'molecule_type': 'Unknown'}
        
        cluster_sheet.cell(row=row_idx, column=1, value=f"C{int(cluster_num):04d}")
        cluster_sheet.cell(row=row_idx, column=2, value=stats.get('Num bases covered', ''))
        cluster_sheet.cell(row=row_idx, column=3, value=f"{stats.get('Frac bases covered', 0):.4f}")
        cluster_sheet.cell(row=row_idx, column=4, value=f"{stats.get('Frac bases covered over unambig', 0):.4f}")
        cluster_sheet.cell(row=row_idx, column=5, value=f"{stats.get('Average coverage/depth', 0):.4f}")
        cluster_sheet.cell(row=row_idx, column=6, value=f"{stats.get('Average coverage/depth over unambig', 0):.4f}")
        cluster_sheet.cell(row=row_idx, column=7, value=taxonomy.get('family', 'Unknown'))
        cluster_sheet.cell(row=row_idx, column=8, value=taxonomy.get('genus', 'Unknown'))
        cluster_sheet.cell(row=row_idx, column=9, value=taxonomy.get('species', 'Unknown'))
        cluster_sheet.cell(row=row_idx, column=10, value=taxonomy.get('molecule_type', 'Unknown'))
        cluster_sheet.cell(row=row_idx, column=11, value=len(cluster_accessions.get(cluster_num, [])))
    
    # Sheet 2: Probe Mapping (with coordinates, excluding adapter portion)
    probe_sheet = workbook.create_sheet('ProbeMapping')
    
    probe_headers = [
        'Probe_ID', 'Cluster', 'Start_Pos', 'End_Pos', 'Probe_Length',
        'Family', 'Genus', 'Species', 'Molecule_type',
        'Original_Sequence', 'Has_Adapter'
    ]
    
    for col, header in enumerate(probe_headers, 1):
        cell = probe_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, (probe_id, details) in enumerate(sorted(probe_details.items(), key=lambda x: x[1].get('cluster_num') or 9999), 2):
        taxonomy = details.get('taxonomy', {})
        
        probe_sheet.cell(row=row_idx, column=1, value=probe_id)
        probe_sheet.cell(row=row_idx, column=2, value=f"C{int(details['cluster_num']):04d}" if details['cluster_num'] is not None else 'Unmapped')
        probe_sheet.cell(row=row_idx, column=3, value=details['start'] if details['start'] is not None else '')
        probe_sheet.cell(row=row_idx, column=4, value=details['end'] if details['end'] is not None else '')
        probe_sheet.cell(row=row_idx, column=5, value=len(details['original_sequence']))
        probe_sheet.cell(row=row_idx, column=6, value=taxonomy.get('family', 'Unknown'))
        probe_sheet.cell(row=row_idx, column=7, value=taxonomy.get('genus', 'Unknown'))
        probe_sheet.cell(row=row_idx, column=8, value=taxonomy.get('species', 'Unknown'))
        probe_sheet.cell(row=row_idx, column=9, value=taxonomy.get('molecule_type', 'Unknown'))
        probe_sheet.cell(row=row_idx, column=10, value=details['original_sequence'])
        
        # Check if probe contains adapter sequence
        has_adapter = adapter_seq in details['original_sequence'] or \
                     adapter_seq[::-1] in details['original_sequence']
        probe_sheet.cell(row=row_idx, column=11, value='Yes' if has_adapter else 'No')
    
    # Sheet 3: Unmapped Probes
    unmapped_sheet = workbook.create_sheet('UnmappedProbes')
    
    unmapped_headers = [
        'Probe_ID', 'Original_Sequence', 'Core_Sequence', 'Num_Sequences_Mapped',
        'Adapter_Left', 'Adapter_Right'
    ]
    
    for col, header in enumerate(unmapped_headers, 1):
        cell = unmapped_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    unmapped_probes = [(pid, pr) for pid, pr in mapping_results.items() if not pr.get('cluster_num')]
    for row_idx, (probe_id, result) in enumerate(unmapped_probes, 2):
        adapter = result['adapter_info']
        unmapped_sheet.cell(row=row_idx, column=1, value=probe_id)
        unmapped_sheet.cell(row=row_idx, column=2, value=result['original_sequence'])
        unmapped_sheet.cell(row=row_idx, column=3, value=result['core_sequence'])
        unmapped_sheet.cell(row=row_idx, column=4, value=result['stats'].get('num_sequences', 0))
        unmapped_sheet.cell(row=row_idx, column=5, value='Yes' if adapter.get('has_adapter_left') else 'No')
        unmapped_sheet.cell(row=row_idx, column=6, value='Yes' if adapter.get('has_adapter_right') else 'No')
    
    # Sheet 4: Taxonomy Summary
    tax_sheet = workbook.create_sheet('TaxonomySummary')
    
    tax_headers = [
        'TaxID', 'Family', 'Genus', 'Species', 'Clusters', 'Num_Probes'
    ]
    
    for col, header in enumerate(tax_headers, 1):
        cell = tax_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Aggregate by taxonomy
    tax_groups = {}
    for cluster_num, tax in cluster_taxonomy.items():
        tax_key = (tax.get('taxid'), tax.get('family'), tax.get('genus'), tax.get('species'))
        if tax_key not in tax_groups:
            tax_groups[tax_key] = []
        tax_groups[tax_key].append(cluster_num)
    
    for row_idx, (taxid, family, genus, species), clusters in enumerate(
        sorted([(k[0], k[1], k[2], k[3], v) for k, v in tax_groups.items()], key=lambda x: x[0] or 999999), 2):
        tax_sheet.cell(row=row_idx, column=1, value=taxid or '')
        tax_sheet.cell(row=row_idx, column=2, value=family)
        tax_sheet.cell(row=row_idx, column=3, value=genus)
        tax_sheet.cell(row=row_idx, column=4, value=species)
        tax_sheet.cell(row=row_idx, column=5, value=', '.join([f"C{int(c):04d}" for c in sorted(clusters, key=int)]))
        tax_sheet.cell(row=row_idx, column=6, value=len([p for p in mapping_results.values() if p.get('cluster_num') in clusters]))
    
    # Adjust column widths
    for sheet in workbook:
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 10 for cell in column)
            adjusted_width = min(max_length + 2, 50)
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output_path)
    print(f"Wrote probe mapping Excel to: {output_path}")


def _write_probe_mapping_csv(output_path: str, 
                             probe_details: Dict[str, Dict],
                             analysis: Dict[str, Dict],
                             cluster_accessions: Dict[str, List[str]],
                             metadata_map: Dict[str, Dict],
                             adapter_seq: str = 'GTGGAGGTCGCTAGATGGTC') -> None:
    """
    Fallback CSV output when openpyxl not available.
    
    Args:
        output_path: output CSV file path
        probe_details: dict mapping probe_id to probe details
        analysis: dict mapping genome_id to analysis stats
        cluster_accessions: dict mapping cluster_num to list of accession IDs
        metadata_map: dict mapping accession_id to taxonomy info
        adapter_seq: adapter sequence
    """
    import csv
    
    with open(output_path, 'w', newline='') as f:
        writer = csv.writer(f)
        
        # Write ClusterSummary sheet
        writer.writerow(['ClusterSummary'])
        writer.writerow([
            'Cluster_ID', 'Num_bases_covered', 'Frac_bases_covered', 
            'Frac_bases_covered_over_unambig', 'Average_coverage_depth',
            'Average_coverage_depth_over_unambig', 'Family', 'Genus', 
            'Species', 'Molecule_type', 'Num_accessions'
        ])
        
        cluster_nums = set(cluster_accessions.keys())
        for cluster_num in sorted(cluster_nums, key=int):
            genome_id = None
            for gid in analysis.keys():
                if f'genome {cluster_num}' in gid:
                    genome_id = gid
                    break
            
            stats = analysis.get(genome_id, {}) if genome_id else {}
            taxonomy = {}
            for acc in cluster_accessions.get(cluster_num, []):
                if acc in metadata_map:
                    taxonomy = metadata_map[acc]
                    break
            
            if not taxonomy:
                taxonomy = {'family': 'Unknown', 'genus': 'Unknown', 'species': 'Unknown', 'molecule_type': 'Unknown'}
            
            writer.writerow([
                f"C{int(cluster_num):04d}",
                stats.get('Num bases covered', ''),
                f"{stats.get('Frac bases covered', 0):.4f}",
                f"{stats.get('Frac bases covered over unambig', 0):.4f}",
                f"{stats.get('Average coverage/depth', 0):.4f}",
                f"{stats.get('Average coverage/depth over unambig', 0):.4f}",
                taxonomy.get('family', 'Unknown'),
                taxonomy.get('genus', 'Unknown'),
                taxonomy.get('species', 'Unknown'),
                taxonomy.get('molecule_type', 'Unknown'),
                len(cluster_accessions.get(cluster_num, []))
            ])
        
        # Write ProbeMapping sheet
        writer.writerow([])
        writer.writerow(['ProbeMapping'])
        writer.writerow([
            'Probe_ID', 'Cluster', 'Start_Pos', 'End_Pos', 'Probe_Length',
            'Family', 'Genus', 'Species', 'Molecule_type',
            'Original_Sequence', 'Has_Adapter'
        ])
        
        for probe_id, details in sorted(probe_details.items(), key=lambda x: x[1].get('cluster_num') or 9999):
            taxonomy = details.get('taxonomy', {})
            has_adapter = adapter_seq in details['original_sequence'] or adapter_seq[::-1] in details['original_sequence']
            
            writer.writerow([
                probe_id,
                f"C{int(details['cluster_num']):04d}" if details['cluster_num'] is not None else 'Unmapped',
                details['start'] if details['start'] is not None else '',
                details['end'] if details['end'] is not None else '',
                len(details['original_sequence']),
                taxonomy.get('family', 'Unknown'),
                taxonomy.get('genus', 'Unknown'),
                taxonomy.get('species', 'Unknown'),
                taxonomy.get('molecule_type', 'Unknown'),
                details['original_sequence'],
                'Yes' if has_adapter else 'No'
            ])
    
    print(f"Wrote probe mapping CSV to: {output_path}")
